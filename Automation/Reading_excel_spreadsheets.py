import openpyxl
import os

os.chdir('C:\\Users\\NM Siddharth\\Downloads')

workbook = openpyxl.load_workbook('example.xlsx')
print(type(workbook))

sheet = workbook['Sheet1']  # gets sheet object
print(type(sheet))

print(workbook.sheetnames)  # Gets list of sheets

cell = sheet['A1']
print(str(cell.value))  # Returns A1 cell content

for i in range(1,8):
    print(i, sheet.cell(row= i, column =2 ).value)

