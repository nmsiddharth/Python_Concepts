import openpyxl

wb = openpyxl.Workbook() # creates new blank EXCEL spreadsheet

print(wb.get_sheet_names()) # returns sheets in list form

sheet = wb.get_sheet_by_name('Sheet')  # sheet object
print(sheet)

print(sheet['A1'].value)   # returns 'None' since no value present

sheet['A1'] = 'Siddu'
sheet['A2'] = 24

import os

os.chdir('C:\\Users\\NM Siddharth\\Desktop')
wb.save('Demo.xlsx')
sheet2 = wb.create_sheet()  # used to create new sheets
print(sheet2)

print(wb.get_sheet_names())
# OR                # used to return list of sheets
print(wb.sheetnames)

sheet2.title = 'My_new_sheet'  # used to change name of sheet

wb.create_sheet(index=0,title='first sheet') # all in one function regarding sheet

print(wb.get_sheet_names())
wb.save('Demo.xlsx')
